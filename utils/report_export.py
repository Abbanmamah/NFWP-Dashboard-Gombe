from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def create_management_report(
    summary_df,
    trend_df,
    dq,
    top_df,
    bottom_df,
    total_wags,
    total_members,
    total_savings,
    total_loan_fund,
    total_loan_disbursed,
    total_loan_repaid,
    outstanding_loan,
    average_utilisation,
):

    wb = Workbook()

    ws = wb.active
    ws.title = "Executive Summary"

    ws["A1"] = "NFWP-SU GOMBE STATE"
    ws["A1"].font = Font(size=18, bold=True)

    ws["A2"] = "Management Dashboard Report"
    ws["A2"].font = Font(size=14, bold=True)

    rows = [

        ("Total WAGs", total_wags),

        ("Total Members", total_members),

        ("Total Savings", total_savings),

        ("Total Loan Fund", total_loan_fund),

        ("Total Loan Disbursed", total_loan_disbursed),

        ("Total Loan Repaid", total_loan_repaid),

        ("Outstanding Loan", outstanding_loan),

        ("Average Loan Utilisation (%)", average_utilisation),

    ]

    start_row = 5

    for title, value in rows:

        ws.cell(row=start_row, column=1).value = title
        ws.cell(row=start_row, column=1).font = Font(bold=True)

        ws.cell(row=start_row, column=2).value = value

        start_row += 1
# ==========================================
    # LGA SUMMARY
    # ==========================================

    ws2 = wb.create_sheet("LGA Summary")

    headers = list(summary_df.columns)

    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    for row_index, row in enumerate(
        summary_df.itertuples(index=False),
        start=2,
    ):
        for col_index, value in enumerate(row, start=1):
            ws2.cell(
                row=row_index,
                column=col_index
            ).value = value

    # ==========================================
    # MONTHLY TRENDS
    # ==========================================

    ws3 = wb.create_sheet("Monthly Trends")

    headers = list(trend_df.columns)

    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    for row_index, row in enumerate(
        trend_df.itertuples(index=False),
        start=2,
    ):
        for col_index, value in enumerate(row, start=1):
            ws3.cell(
                row=row_index,
                column=col_index
            ).value = value
            # ==========================================
    # DATA QUALITY
    # ==========================================

    ws4 = wb.create_sheet("Data Quality")

    ws4["A1"] = "Data Quality Dashboard"
    ws4["A1"].font = Font(size=14, bold=True)

    row = 3

    for key, value in dq.items():

        ws4.cell(row=row, column=1).value = key
        ws4.cell(row=row, column=1).font = Font(bold=True)

        ws4.cell(row=row, column=2).value = value

        row += 1

    # ==========================================
    # TOP & BOTTOM WAGS
    # ==========================================

    ws5 = wb.create_sheet("Top & Bottom WAGs")

    ws5["A1"] = "Top 10 WAGs"
    ws5["A1"].font = Font(size=14, bold=True)

    headers = list(top_df.columns)

    for col, header in enumerate(headers, start=1):

        cell = ws5.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    for r, row_data in enumerate(
        top_df.itertuples(index=False),
        start=3,
    ):

        for c, value in enumerate(row_data, start=1):

            ws5.cell(row=r, column=c).value = value

    start = len(top_df) + 6

    ws5.cell(
        row=start,
        column=1
    ).value = "Bottom 10 WAGs"

    ws5.cell(
        row=start,
        column=1
    ).font = Font(size=14, bold=True)

    headers = list(bottom_df.columns)

    for col, header in enumerate(headers, start=1):

        cell = ws5.cell(
            row=start + 1,
            column=col
        )

        cell.value = header
        cell.font = Font(bold=True)

    for r, row_data in enumerate(
        bottom_df.itertuples(index=False),
        start=start + 2,
    ):

        for c, value in enumerate(row_data, start=1):

            ws5.cell(
                row=r,
                column=c
            ).value = value

    # ==========================================
    # SAVE REPORT
    # ==========================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output.getvalue()
    